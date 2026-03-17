"""
Biotech Prospectus Risk Factor Analyzer v5.2
=============================================
Three-pass approach:
  Pass 1: Line-level font detection (formatted PDFs)
  Pass 2: Indentation-based paragraph detection (unformatted PDFs)
  Pass 3: Text pattern fallback

Requirements: pip install PyMuPDF openpyxl
"""

import os, re, json, csv, fitz
from pathlib import Path
from dataclasses import dataclass, field, asdict
from typing import Optional
from collections import Counter

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

# put in the path to your PDFs here
PDF_FOLDER = "#insert pdf file path"
OUTPUT_CSV = "./results.csv"
DEBUG = True  # Set to True for diagnostics

@dataclass
class Span:
    text: str; font_name: str; font_size: float
    is_bold: bool; is_italic: bool; page_num: int
    y_pos: float; x_pos: float

@dataclass
class Line:
    spans: list; text: str; page_num: int; y_pos: float
    is_bold: bool; is_italic: bool; is_all_caps: bool
    font_size: float; char_count: int
    gap_before: float = 0.0

@dataclass
class Paragraph:
    lines: list; text: str; first_line_bold: bool; all_bold: bool
    all_italic: bool; word_count: int; char_count: int; is_short: bool

@dataclass
class RiskFactor:
    title: str; body: str; word_count: int

@dataclass
class DocumentStyle:
    body_font_size: float; body_font_name: str; body_is_bold: bool
    avg_line_gap: float; avg_para_gap: float

@dataclass
class ProspectusAnalysis:
    filename: str; total_word_count: int; risk_factors_section_word_count: int
    num_risk_factors: int; risk_factors: list = field(default_factory=list)
    extraction_method: str = ""; warnings: list = field(default_factory=list)

# ============================================================================
# STEP 1: EXTRACT SPANS
# ============================================================================
def extract_spans(pdf_path):
    doc = fitz.open(pdf_path)
    all_spans, plain = [], []
    for pn, page in enumerate(doc):
        for block in page.get_text("dict", flags=fitz.TEXT_PRESERVE_WHITESPACE)["blocks"]:
            if block["type"] != 0: continue
            for line in block["lines"]:
                lt = []
                for s in line["spans"]:
                    t = s["text"]
                    if not t.strip(): lt.append(t); continue
                    fn, fs, fl, bb = s["font"], round(s["size"],1), s["flags"], s["bbox"]
                    ib = bool(fl&(1<<4)) or any(x in fn.lower() for x in ["bold","-bd","heavy","black","demi","semibold"])
                    ii = bool(fl&(1<<1)) or any(x in fn.lower() for x in ["italic","oblique","slant","-it"])
                    all_spans.append(Span(t,fn,fs,ib,ii,pn,round(bb[1],1),round(bb[0],1)))
                    lt.append(t)
                plain.append("".join(lt))
    doc.close()
    return "\n".join(plain), all_spans

# ============================================================================
# STEP 2: BUILD LINES WITH GAP INFO
# ============================================================================
def build_lines(spans):
    if not spans: return []
    lines, cur = [], [spans[0]]
    cp, cy = spans[0].page_num, spans[0].y_pos
    for s in spans[1:]:
        if s.page_num == cp and abs(s.y_pos - cy) < 2.0:
            cur.append(s)
        else:
            ln = _fin_line(cur, cp)
            if ln: lines.append(ln)
            cur, cp, cy = [s], s.page_num, s.y_pos
    ln = _fin_line(cur, cp)
    if ln: lines.append(ln)
    for i in range(1, len(lines)):
        if lines[i].page_num == lines[i-1].page_num:
            lines[i].gap_before = lines[i].y_pos - lines[i-1].y_pos
        else:
            lines[i].gap_before = 999
    return lines

def _fin_line(spans, pn):
    t = "".join(s.text for s in spans).strip()
    if not t: return None
    bc = sum(len(s.text.strip()) for s in spans if s.is_bold)
    ic = sum(len(s.text.strip()) for s in spans if s.is_italic)
    tc = sum(len(s.text.strip()) for s in spans)
    if tc == 0: return None
    sc = Counter()
    for s in spans:
        n = len(s.text.strip())
        if n > 0: sc[s.font_size] += n
    ds = sc.most_common(1)[0][0]
    return Line(spans=spans, text=t, page_num=pn, y_pos=spans[0].y_pos,
                is_bold=bc>tc*0.6, is_italic=ic>tc*0.6,
                is_all_caps=t==t.upper() and len(t)>5 and any(c.isalpha() for c in t),
                font_size=ds, char_count=tc)

# ============================================================================
# STEP 3: LEARN DOCUMENT STYLE
# ============================================================================
def learn_style(lines):
    sz, fn = Counter(), Counter()
    bc, tc = 0, 0
    gaps = []
    for ln in lines:
        if ln.char_count > 40:
            sz[ln.font_size] += ln.char_count
            for s in ln.spans: fn[s.font_name] += len(s.text.strip())
            if ln.is_bold: bc += ln.char_count
            tc += ln.char_count
        if 0 < ln.gap_before < 100: gaps.append(ln.gap_before)
    avg_gap = sum(gaps)/len(gaps) if gaps else 12.0
    sorted_gaps = sorted(gaps)
    if len(sorted_gaps) > 10:
        small = [g for g in sorted_gaps if g < avg_gap * 1.5]
        large = [g for g in sorted_gaps if g >= avg_gap * 1.5]
        lg = sum(small)/len(small) if small else avg_gap
        pg = sum(large)/len(large) if large else avg_gap * 2
    else:
        lg, pg = avg_gap, avg_gap * 2
    return DocumentStyle(
        body_font_size=sz.most_common(1)[0][0] if sz else 10.0,
        body_font_name=fn.most_common(1)[0][0] if fn else "",
        body_is_bold=bc > tc * 0.5 if tc else False,
        avg_line_gap=lg, avg_para_gap=pg)

# ============================================================================
# STEP 4: BUILD PARAGRAPHS
# ============================================================================
def build_paragraphs(lines, style):
    if not lines: return []
    gap_threshold = style.avg_line_gap * 1.4
    if gap_threshold < 2: gap_threshold = style.avg_line_gap + 2
    paras, current = [], [lines[0]]
    for ln in lines[1:]:
        if ln.gap_before > gap_threshold or ln.gap_before >= 999:
            p = _fin_para(current)
            if p: paras.append(p)
            current = [ln]
        else:
            current.append(ln)
    p = _fin_para(current)
    if p: paras.append(p)
    return paras

def _fin_para(lines):
    text = " ".join(ln.text.strip() for ln in lines if ln.text.strip())
    if not text: return None
    wc = len(text.split())
    cc = sum(ln.char_count for ln in lines)
    ab = all(ln.is_bold for ln in lines if ln.text.strip())
    ai = all(ln.is_italic for ln in lines if ln.text.strip())
    fb = lines[0].is_bold if lines else False
    return Paragraph(lines=lines, text=text, first_line_bold=fb,
                     all_bold=ab, all_italic=ai, word_count=wc,
                     char_count=cc, is_short=wc<=45 and cc<500)

# ============================================================================
# STEP 5: FIND RISK FACTORS SECTION
# ============================================================================
def find_risk_section(lines):
    end_pats = [
        r"use\s+of\s+proceeds", r"cautionary\s+note.*",
        r"special\s+note\s+regarding\s+forward.*", r"note\s+regarding\s+forward.*",
        r"forward[- ]looking\s+statements",
        r"dividend\s+policy", r"capitalization", r"dilution",
        r"market\s+(?:and\s+)?industry\s+data", r"industry\s+and\s+market\s+data",
        r"selected\s+(?:consolidated\s+)?financial", r"management[''\u2019]?s?\s+discussion",
        r"determination\s+of\s+(?:the\s+)?offering", r"plan\s+of\s+distribution",
        r"description\s+of\s+(?:share\s+)?capital",
        r"description\s+of\s+(?:the\s+)?(?:notes|securities|debt)",
        r"material\s+(?:u\.?s\.?\s+)?(?:federal\s+)?(?:income\s+)?tax",
        r"underwriting\b", r"legal\s+matters", r"experts?\s*$",
        r"where\s+you\s+can\s+find", r"prospectus\s+summary",
        r"notes?\s+to\s+(?:consolidated\s+)?financial",
    ]
    cands = []
    for i, ln in enumerate(lines):
        tc = re.sub(r'\s+',' ',ln.text.strip())
        if re.match(r'^risk\s+factors\s*$', tc, re.IGNORECASE):
            sc = 0
            if ln.is_bold: sc += 3
            if ln.is_all_caps: sc += 2
            if ln.font_size > 10: sc += 2
            if ln.char_count < 30: sc += 1
            toc = False
            if i>0 and i<len(lines)-1:
                nt = lines[i+1].text.strip()
                if re.match(r'^[SsFf]?-?\d{1,3}$', nt): toc = True
                pt = lines[i-1].text.strip()
                if len(pt)<60 and len(nt)<60 and not ln.is_bold and ln.font_size<=10: toc = True
            if not toc: sc += 2
            cands.append((i, sc, toc))
    if not cands:
        for i, ln in enumerate(lines):
            tc = re.sub(r'\s+',' ',ln.text.strip())
            if re.match(r'^risk\s+factors', tc, re.IGNORECASE) and ln.char_count < 100:
                if ln.is_bold or ln.is_all_caps or ln.font_size > 10:
                    cands.append((i, 3, False))
    if not cands: return None
    cands.sort(key=lambda x: (-x[1], x[2], x[0]))
    si = cands[0][0] + 1
    if DEBUG: print(f"    RF heading at line {cands[0][0]}: '{lines[cands[0][0]].text.strip()}'")
    for i in range(si+3, len(lines)):
        tc = re.sub(r'\s+',' ',lines[i].text.strip())
        for pat in end_pats:
            if re.match(rf'^\s*{pat}\s*$', tc, re.IGNORECASE):
                ih = (lines[i].is_bold or lines[i].is_all_caps or lines[i].font_size>10
                      or (len(tc.split())<=6 and lines[i].char_count<80))
                if ih:
                    if DEBUG: print(f"    RF section ends at line {i}: '{tc[:80]}'")
                    return (si, i)
    return (si, len(lines))

# ============================================================================
# STEP 6: SEGMENT RISK FACTORS (3-pass)
# ============================================================================
INTRO_PAT = re.compile(
    r'(investing in .{0,20} involves|you should carefully consider|'
    r'if any of the following risks|could decline and you could lose|'
    r'risks and uncertainties .{0,30} not presently known|'
    r'additional risks and uncertainties|before deciding to invest|'
    r'the risks .{0,20} described below are not the only|'
    r'an investment in .{0,20} offered|the occurrence of any|'
    r'in that case|the risks and uncertainties described|'
    r'not available or are not available on acceptable terms|'
    r'should we require additional|if additional funds are required)',
    re.IGNORECASE)

RF_STARTERS = re.compile(
    r'^(We\s|Our\s|If\s|The\s(?!risks?\s|following|extent\s|pharmaceutical)|There\s|Any\s|'
    r'A\s(?!high|large|significant\s+portion)|An\s|Changes\s|Loss\s|Failure\s|'
    r'Inability\s|You\s(?!should)|Certain\s|Future\s|Adverse\s|Regulatory\s|'
    r'Competition|Market\s|Economic\s|Clinical\s|Product\s|Intellectual\s|'
    r'Securities\s|Anti-|Data\s|Cyber|Provisions\s|Federal\s|State\s|'
    r'Substantial\s|Significant\s|Because\s|Although\s|Even\s|Sales\s|'
    r'Raising\s|Coverage\s|Price\s|Confidentiality\s|Third\s|Companies\s|'
    r'New\s|It\s+is\s|Risks\s|Obtaining\s)', re.IGNORECASE)

SUBCAT_PAT = re.compile(
    r'^risks?\s+(related|relating|associated|concerning)\s+(to|with)\s', re.IGNORECASE)

def _looks_like_intro(text):
    """Check if a heading's text looks like introductory/boilerplate text rather than an actual RF title."""
    return bool(INTRO_PAT.search(text))

def _is_noun_phrase_heading(text):
    """Check if text looks like a noun-phrase style heading (common in older prospectuses).
    E.g. 'UNCERTAINTY ASSOCIATED WITH XENOMOUSE TECHNOLOGY'
         'NO ASSURANCE OF SUCCESSFUL PRODUCT DEVELOPMENT'
         'DEPENDENCE ON COLLABORATIVE ARRANGEMENTS'
    These are ALL CAPS, relatively short, and don't read like a sentence."""
    t = text.strip()
    wc = len(t.split())
    if wc < 2 or wc > 25:
        return False
    # Must be all caps
    if t != t.upper():
        return False
    # Should not end with a period (sentences do, headings usually don't)
    if t.endswith('.'):
        return False
    # Should contain mostly alphabetic words (not numbers/tables)
    alpha_words = sum(1 for w in t.split() if any(c.isalpha() for c in w))
    if alpha_words < wc * 0.7:
        return False
    return True


def classify_para(para, style):
    t = para.text.strip()
    wc = para.word_count
    if wc < 2: return 'skip'
    if re.match(r'^(table of contents|page\s+\d|\d+\s*$)', t, re.IGNORECASE): return 'skip'
    if SUBCAT_PAT.match(t) and wc < 20: return 'subheading'
    if re.match(r'^(general\s+risk|additional\s+risk)', t, re.IGNORECASE) and wc < 12: return 'subheading'
    if INTRO_PAT.search(t): return 'intro'
    if para.is_short and wc >= 3:
        if para.all_bold and not style.body_is_bold: return 'heading'
        if para.all_italic: return 'heading'
        if all(ln.is_all_caps for ln in para.lines if ln.text.strip()) and wc <= 40: return 'heading'
        if para.lines[0].font_size > style.body_font_size + 0.5: return 'heading'
    if para.is_short and wc >= 3 and wc <= 45:
        # Filter out bullet points / list items (start with . or ; or bullet chars)
        if re.match(r'^[\.\;\•\-\*]\s', t):
            return 'body'
        # Filter out continuation sentences (start with lowercase)
        if t[0].islower():
            return 'body'
        if RF_STARTERS.match(t) and not INTRO_PAT.search(t):
            return 'heading_candidate'
        if t.rstrip().endswith('.') and wc <= 35 and wc >= 5 and t[0].isupper():
            return 'heading_candidate'
    return 'body'

def segment_risk_factors(lines, start_idx, end_idx, style):
    section_lines = lines[start_idx:end_idx]

    # ===== PASS 1: LINE-LEVEL FONT DETECTION =====
    line_classes = []
    for ln in section_lines:
        t = ln.text.strip()
        wc = len(t.split())
        if wc < 2:
            line_classes.append('skip')
        elif re.match(r'^(table of contents|page\s+\d|\d+\s*$)', t, re.IGNORECASE):
            line_classes.append('skip')
        elif SUBCAT_PAT.match(t) and wc < 20 and (ln.is_bold or ln.is_italic or ln.is_all_caps):
            line_classes.append('subheading')
        elif ln.char_count <= 600 and wc >= 3 and (
            (ln.is_bold and not style.body_is_bold) or ln.is_italic or
            (ln.is_all_caps and wc <= 40) or ln.font_size > style.body_font_size + 0.5):
            line_classes.append('heading')
        else:
            line_classes.append('body')

    font_heading_count = line_classes.count('heading')

    if font_heading_count >= 3:
        if DEBUG: print(f"    Line-level font headings: {font_heading_count} — using font method")
        
        # Debug: show all detected heading lines
        if DEBUG:
            for i, (ln, cls) in enumerate(zip(section_lines, line_classes)):
                if cls == 'heading':
                    print(f"      H-line {i}: [{ln.text.strip()[:100]}]")
        
        merged = []
        i = 0
        while i < len(section_lines):
            cls = line_classes[i]; ln = section_lines[i]
            if cls == 'heading':
                parts = [ln.text.strip()]
                j = i + 1
                while j < len(section_lines) and line_classes[j] == 'heading':
                    parts.append(section_lines[j].text.strip()); j += 1
                merged.append(('heading', " ".join(parts))); i = j
            elif cls == 'subheading':
                parts = [ln.text.strip()]
                j = i + 1
                while j < len(section_lines) and line_classes[j] == 'subheading':
                    parts.append(section_lines[j].text.strip()); j += 1
                merged.append(('subheading', " ".join(parts))); i = j
            elif cls == 'body':
                merged.append(('body', ln.text.strip())); i += 1
            else:
                merged.append(('skip', '')); i += 1

        # --- IMPROVED INTRO DETECTION ---
        # Only demote headings that look like intro paragraphs (long sentences, 
        # contain intro language). Do NOT demote short ALL-CAPS noun-phrase headings
        # or headings that match RF_STARTERS, as these are real risk factor titles.
        #
        # Strategy: walk from the top; once we see either a subheading or a heading 
        # that is clearly a risk-factor title (noun-phrase ALL CAPS, or RF_STARTERS match),
        # stop demoting. Only demote headings before that point IF they look like intro text.
        
        first_real = None
        for idx, (tp, txt) in enumerate(merged):
            if tp == 'subheading':
                first_real = idx
                break
            if tp == 'heading':
                # Is this a real RF title or intro text?
                if _is_noun_phrase_heading(txt):
                    # This is a real heading like "UNCERTAINTY ASSOCIATED WITH XENOMOUSE TECHNOLOGY"
                    first_real = idx
                    break
                if RF_STARTERS.match(txt) and not _looks_like_intro(txt):
                    first_real = idx
                    break
                # If it's long and contains intro language, it's probably intro
                if _looks_like_intro(txt):
                    continue  # skip this, keep looking
                # Short ALL CAPS text that doesn't match intro — likely a heading
                if txt == txt.upper() and len(txt.split()) <= 15:
                    first_real = idx
                    break
                # Default: if it's short enough and doesn't look like intro, treat as real
                if len(txt.split()) <= 25:
                    first_real = idx
                    break
        
        if first_real and first_real > 0:
            if DEBUG:
                print(f"    Demoting {first_real} intro heading(s) before first real heading")
            for idx in range(first_real):
                if merged[idx][0] == 'heading':
                    if DEBUG:
                        print(f"      Demoted: [{merged[idx][1][:80]}]")
                    merged[idx] = ('body', merged[idx][1])

        # --- FILTER OUT FALSE HEADINGS (ALL-CAPS-only documents) ---
        # In monospaced/plain-text PDFs where headings are detected ONLY via ALL CAPS
        # (not bold/italic/font-size), long ALL CAPS body text can be misdetected.
        # For documents with real font formatting (bold/italic/larger size), we trust
        # the font detection completely and do NOT filter by length.
        
        has_real_font_headings = any(
            (ln.is_bold or ln.is_italic or ln.font_size > style.body_font_size + 0.5)
            and not ln.is_all_caps  # must have real formatting beyond just ALL CAPS
            for ln, cls in zip(section_lines, line_classes) if cls == 'heading'
        )
        
        if not has_real_font_headings:
            # ALL-CAPS-only detection: filter out overly long "headings"
            for idx, (tp, txt) in enumerate(merged):
                if tp != 'heading':
                    continue
                wc = len(txt.split())
                if wc > 30 and not _is_noun_phrase_heading(txt):
                    merged[idx] = ('body', txt)
                    if DEBUG:
                        print(f"    Demoted long ALL-CAPS non-heading ({wc}w): [{txt[:80]}]")
        elif DEBUG:
            print(f"    Real font headings detected — skipping length filter")

        rfs, cur_title, body_parts = [], None, []
        for tp, txt in merged:
            if tp == 'heading':
                if cur_title:
                    body = "\n".join(p for p in body_parts if p)
                    wc = len(body.split())
                    if wc >= 3: rfs.append(RiskFactor(cur_title, body, wc))
                cur_title = txt; body_parts = []
            elif tp in ('subheading', 'skip'): continue
            else:
                if txt: body_parts.append(txt)
        if cur_title:
            body = "\n".join(p for p in body_parts if p)
            wc = len(body.split())
            if wc >= 3: rfs.append(RiskFactor(cur_title, body, wc))

        if DEBUG:
            print(f"    After merging & filtering: {len(rfs)} risk factors")
            for rf in rfs[:8]: print(f"      [{rf.word_count:>4}w] {rf.title[:100]}")
            if len(rfs) > 8: print(f"      ... and {len(rfs)-8} more")
        return rfs, "font-standalone-merged"

    # ===== PASS 2: INDENTATION-BASED DETECTION =====
    if DEBUG: print(f"    Font headings: {font_heading_count} (too few) — trying indent method")

    paras = build_paragraphs(section_lines, style)

    # DIAGNOSTIC: Show indent info for first 20 paragraphs
    if DEBUG:
        print(f"    --- INDENT DIAGNOSTIC (first 20 paragraphs) ---")
        count = 0
        for p in paras:
            if p.word_count >= 3 and p.lines and p.lines[0].spans:
                x = p.lines[0].spans[0].x_pos
                print(f"    x={x:>6.1f}  wc={p.word_count:>3}  short={p.is_short}  [{p.text[:75]}]")
                count += 1
                if count >= 20: break
        print(f"    --- END INDENT DIAGNOSTIC ---")

    # Analyze indentation levels
    first_line_xpos = []
    for p in paras:
        if p.word_count >= 3 and p.lines and p.lines[0].spans:
            x = p.lines[0].spans[0].x_pos
            first_line_xpos.append(x)

    x_counter = Counter()
    for x in first_line_xpos:
        x_counter[round(x / 2) * 2] += 1

    indent_method_viable = False
    title_x, body_x = None, None

    if len(x_counter) >= 2:
        top_two = x_counter.most_common(2)
        x1, count1 = top_two[0]
        x2, count2 = top_two[1]
        if abs(x1 - x2) >= 3:
            title_x = min(x1, x2)
            body_x = max(x1, x2)
            indent_method_viable = True
            if DEBUG:
                print(f"    Indent levels: title_x={title_x} ({x_counter[title_x]} paras), "
                      f"body_x={body_x} ({x_counter[body_x]} paras)")
                print(f"    All x levels: {x_counter.most_common(10)}")

    if indent_method_viable:
        threshold = (title_x + body_x) / 2
        rfs, cur_title, body_parts = [], None, []

        for p in paras:
            t = p.text.strip(); wc = p.word_count
            if wc < 2: continue
            if re.match(r'^(table of contents|page\s+\d|\d+\s*$)', t, re.IGNORECASE): continue

            first_x = p.lines[0].spans[0].x_pos if p.lines and p.lines[0].spans else body_x
            first_x_rounded = round(first_x / 2) * 2
            is_title_indent = first_x_rounded <= threshold

            if SUBCAT_PAT.match(t) and wc < 20: continue
            if re.match(r'^(general\s+risk|additional\s+risk|risks\s+related\s+to\s+this)',
                        t, re.IGNORECASE) and wc < 15: continue
            if INTRO_PAT.search(t) and cur_title is None: continue

            if is_title_indent and p.is_short and wc >= 3:
                if cur_title is not None:
                    body = " ".join(body_parts)
                    bwc = len(body.split())
                    if bwc >= 3: rfs.append(RiskFactor(cur_title, body, bwc))
                cur_title = t; body_parts = []
            else:
                if t: body_parts.append(t)

        if cur_title:
            body = " ".join(body_parts)
            bwc = len(body.split())
            if bwc >= 3: rfs.append(RiskFactor(cur_title, body, bwc))

        if len(rfs) >= 3:
            if DEBUG:
                print(f"    Indent method found {len(rfs)} risk factors")
                for rf in rfs[:8]: print(f"      [{rf.word_count:>4}w] {rf.title[:100]}")
                if len(rfs) > 8: print(f"      ... and {len(rfs)-8} more")
            return rfs, "indent-paragraph"
        elif DEBUG:
            print(f"    Indent method found only {len(rfs)} RFs — falling through")

    # ===== PASS 3: TEXT PATTERN FALLBACK =====
    if DEBUG: print(f"    Using text pattern fallback")
    classes = [classify_para(p, style) for p in paras]
    
    # Debug: show all heading_candidates
    if DEBUG:
        for i, (p, c) in enumerate(zip(paras, classes)):
            if c == 'heading_candidate':
                print(f"    CANDIDATE para {i}: wc={p.word_count} [{p.text[:100]}]")
    
    text_candidates = sum(1 for c in classes if c == 'heading_candidate')

    if text_candidates >= 3:
        # Validate candidates: a real heading should be followed by 
        # substantial body content before the next heading
        validated_classes = list(classes)
        for i, c in enumerate(validated_classes):
            if c == 'heading_candidate':
                t = paras[i].text.strip()
                # Double-check: must actually match RF_STARTERS
                if not RF_STARTERS.match(t):
                    validated_classes[i] = 'body'
                    continue
                # Look ahead: find body content before next heading_candidate
                body_words = 0
                has_substantial_body = False
                for j in range(i+1, len(paras)):
                    if validated_classes[j] == 'heading_candidate':
                        break
                    if validated_classes[j] in ('body', 'intro'):
                        wc_j = paras[j].word_count
                        body_words += wc_j
                        if wc_j > 30:
                            has_substantial_body = True
                            break
                        if body_words > 50:
                            has_substantial_body = True
                            break
                if has_substantial_body:
                    validated_classes[i] = 'heading'
                else:
                    validated_classes[i] = 'body'
        classes = validated_classes
        method = "text-paragraph"
    else:
        for i, (p, c) in enumerate(zip(paras, classes)):
            if c == 'body' and p.is_short and p.word_count >= 5 and RF_STARTERS.match(p.text):
                classes[i] = 'heading'
        method = "fallback-paragraph"

    first_heading = None
    for i, c in enumerate(classes):
        if c in ('heading', 'subheading'): first_heading = i; break
    if first_heading and first_heading > 0:
        for i in range(first_heading):
            if classes[i] in ('heading', 'heading_candidate'): classes[i] = 'intro'

    rfs, cur_title, body_parts = [], None, []
    for para, cls in zip(paras, classes):
        if cls == 'heading':
            if cur_title:
                body = " ".join(body_parts)
                wc = len(body.split())
                if wc >= 3: rfs.append(RiskFactor(cur_title, body, wc))
            cur_title = para.text.strip(); body_parts = []
        elif cls in ('subheading', 'skip', 'intro'): continue
        else:
            if para.text.strip(): body_parts.append(para.text.strip())
    if cur_title:
        body = " ".join(body_parts)
        wc = len(body.split())
        if wc >= 3: rfs.append(RiskFactor(cur_title, body, wc))

    return rfs, method

# ============================================================================
# VALIDATION
# ============================================================================
def validate(a):
    w = []
    if a.num_risk_factors == 0: w.append("No risk factors found."); return w
    if a.num_risk_factors < 5: w.append(f"Only {a.num_risk_factors} risk factors.")
    if a.num_risk_factors > 120: w.append(f"{a.num_risk_factors} RFs — over-splitting?")
    wcs = [rf.word_count for rf in a.risk_factors]
    if wcs:
        med = sorted(wcs)[len(wcs)//2]
        if med < 20: w.append(f"Median {med} words/factor — over-detected?")
        if med > 1000: w.append(f"Median {med} words/factor — under-detected?")
    return w

# ============================================================================
# MAIN PIPELINE
# ============================================================================
def analyze_prospectus(pdf_path):
    fn = Path(pdf_path).name
    print(f"\nAnalyzing: {fn}")
    print("-" * 60)
    print("  Extracting text...")
    pt, spans = extract_spans(pdf_path)
    tw = len(pt.split())
    print(f"  Total word count: {tw:,}")
    if tw == 0:
        return ProspectusAnalysis(fn, 0, 0, 0, extraction_method="FAILED-NO-TEXT",
                                  warnings=["No text — may need OCR"])
    print("  Building lines...")
    lines = build_lines(spans)
    print(f"  Found {len(lines)} lines")
    style = learn_style(lines)
    print(f"  Body: {style.body_font_name} @ {style.body_font_size}pt (bold={style.body_is_bold})")
    print(f"  Gaps: line={style.avg_line_gap:.1f}, para={style.avg_para_gap:.1f}")
    print("  Locating Risk Factors section...")
    sr = find_risk_section(lines)
    if sr is None:
        return ProspectusAnalysis(fn, tw, 0, 0, extraction_method="FAILED-NO-SECTION",
                                  warnings=["Could not locate Risk Factors section"])
    si, ei = sr
    st = " ".join(ln.text for ln in lines[si:ei])
    rw = len(st.split())
    print(f"  Risk Factors: ~{rw:,} words (lines {si}-{ei})")
    print("  Segmenting risk factors...")
    rfs, method = segment_risk_factors(lines, si, ei, style)
    print(f"  Found {len(rfs)} risk factors (method: {method})")
    a = ProspectusAnalysis(fn, tw, rw, len(rfs), rfs, method)
    a.warnings = validate(a)
    for w in a.warnings: print(f"  WARNING: {w}")
    if DEBUG and rfs:
        wcs = [rf.word_count for rf in rfs]
        print(f"\n    --- DIAGNOSTICS ---")
        print(f"    Word counts: min={min(wcs)}, max={max(wcs)}, "
              f"median={sorted(wcs)[len(wcs)//2]}, total={sum(wcs)}")
        print(f"    First 10 titles:")
        for rf in rfs[:10]: print(f"      [{rf.word_count:>4}w] {rf.title[:120]}")
        if len(rfs) > 10:
            print(f"    ... ({len(rfs)-13} more) ...")
            print(f"    Last 3 titles:")
            for rf in rfs[-3:]: print(f"      [{rf.word_count:>4}w] {rf.title[:120]}")
        print(f"    --- END DIAGNOSTICS ---\n")
    return a

# ============================================================================
# OUTPUT
# ============================================================================
def save_results(analyses, op):
    sp = op.replace(".csv","_summary.csv")
    with open(sp,'w',newline='',encoding='utf-8') as f:
        w = csv.writer(f)
        w.writerow(["Filename","Total Words","Risk Section Words","Num Risk Factors",
                     "Avg Words/Factor","Min","Max","Median","Method","Warnings"])
        for a in analyses:
            wcs = [rf.word_count for rf in a.risk_factors]
            avg = round(sum(wcs)/len(wcs)) if wcs else 0
            mn,mx,med = (min(wcs),max(wcs),sorted(wcs)[len(wcs)//2]) if wcs else (0,0,0)
            w.writerow([a.filename,a.total_word_count,a.risk_factors_section_word_count,
                        a.num_risk_factors,avg,mn,mx,med,a.extraction_method,
                        "; ".join(a.warnings) if a.warnings else ""])
    print(f"\nSummary:  {sp}")

    dp = op.replace(".csv","_detailed.csv")
    with open(dp,'w',newline='',encoding='utf-8') as f:
        w = csv.writer(f)
        w.writerow(["Filename","RF #","Risk Factor Title","Explanation Word Count"])
        for ai, a in enumerate(analyses):
            if ai > 0: w.writerow([])
            for i,rf in enumerate(a.risk_factors,1):
                w.writerow([a.filename,i,rf.title,rf.word_count])
    print(f"Detailed: {dp}")

    jp = op.replace(".csv","_full.json")
    with open(jp,'w',encoding='utf-8') as f:
        data = []
        for a in analyses:
            d = asdict(a)
            for rf in d.get("risk_factors",[]): rf.pop("spans",None)
            data.append(d)
        json.dump(data,f,indent=2,ensure_ascii=False,default=str)
    print(f"JSON:     {jp}")

    if HAS_OPENPYXL: save_excel(analyses, op)
    else: print("\n  pip install openpyxl for Excel")

def save_excel(analyses, op):
    ep = op.replace(".csv",".xlsx")
    wb = Workbook()
    hf = PatternFill(start_color="4472C4",end_color="4472C4",fill_type="solid")
    hfn = Font(bold=True,size=11,color="FFFFFF")
    wr = Alignment(wrap_text=True,vertical="top")
    ta = Alignment(vertical="top")
    tb = Border(bottom=Side(style="thin",color="D9D9D9"))
    wf = Font(color="CC0000")

    ws = wb.active; ws.title = "Summary"
    for col,h in enumerate(["Filename","Total Words","Risk Section Words",
        "# Risk Factors","Avg Words/Factor","Min","Max","Median","Method","Warnings"],1):
        c = ws.cell(row=1,column=col,value=h); c.font=hfn; c.fill=hf
    for row,a in enumerate(analyses,2):
        wcs = [rf.word_count for rf in a.risk_factors]
        avg = round(sum(wcs)/len(wcs)) if wcs else 0
        mn,mx,med = (min(wcs),max(wcs),sorted(wcs)[len(wcs)//2]) if wcs else (0,0,0)
        wt = "; ".join(a.warnings) if a.warnings else ""
        for col,val in enumerate([a.filename,a.total_word_count,a.risk_factors_section_word_count,
            a.num_risk_factors,avg,mn,mx,med,a.extraction_method,wt],1):
            c = ws.cell(row=row,column=col,value=val); c.border=tb; c.alignment=ta
            if col==10 and wt: c.font=wf
    for i,w in enumerate([40,12,16,14,16,10,10,10,20,50],1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A2"

    ws2 = wb.create_sheet("Risk Factors Detail")
    for col,h in enumerate(["Filename","RF #","Risk Factor Title","Explanation Word Count"],1):
        c = ws2.cell(row=1,column=col,value=h); c.font=hfn; c.fill=hf
    row = 2
    for ai, a in enumerate(analyses):
        if ai > 0:
            for col in range(1, 5):
                ws2.cell(row=row, column=col).fill = PatternFill(
                    start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
            row += 1
        for i,rf in enumerate(a.risk_factors,1):
            for col,val in enumerate([a.filename,i,rf.title,rf.word_count],1):
                c = ws2.cell(row=row,column=col,value=val)
                c.border=tb; c.alignment=wr if col==3 else ta
            row += 1
    for i,w in enumerate([35,8,70,22],1):
        ws2.column_dimensions[get_column_letter(i)].width = w
    ws2.freeze_panes = "A2"
    wb.save(ep); print(f"Excel:    {ep}")

def main():
    pf = Path(PDF_FOLDER)
    if not pf.exists(): print(f"ERROR: '{PDF_FOLDER}' not found."); return
    pdfs = sorted(pf.glob("*.pdf"))
    if not pdfs: print(f"No PDFs in '{PDF_FOLDER}'."); return
    print(f"Found {len(pdfs)} PDF(s)")
    print(f"Debug: {'ON' if DEBUG else 'OFF'}"); print("="*60)
    analyses = []
    for p in pdfs:
        try: analyses.append(analyze_prospectus(str(p)))
        except Exception as e:
            print(f"  ERROR {p.name}: {e}")
            import traceback; traceback.print_exc()
    if analyses:
        save_results(analyses, OUTPUT_CSV)
        print("\n"+"="*90); print("SUMMARY"); print("="*90)
        hdr = f"{'File':<40} {'Words':>8} {'RFs':>5} {'Avg':>6} {'Med':>6} {'Method':<20} {'Warn':>4}"
        print(hdr); print("-"*90)
        for a in analyses:
            wcs = [rf.word_count for rf in a.risk_factors]
            avg = round(sum(wcs)/len(wcs)) if wcs else 0
            med = sorted(wcs)[len(wcs)//2] if wcs else 0
            print(f"{a.filename:<40} {a.total_word_count:>8,} {a.num_risk_factors:>5} "
                  f"{avg:>6} {med:>6} {a.extraction_method:<20} {len(a.warnings):>4}")
        warned = [a for a in analyses if a.warnings]
        if warned:
            print(f"\n  Warnings:")
            for a in warned:
                for w in a.warnings: print(f"    {a.filename}: {w}")

if __name__ == "__main__":
    main()